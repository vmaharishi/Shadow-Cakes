from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
from io import BytesIO, StringIO
import openpyxl
import csv

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== FILE PARSING HELPER ====================

def parse_uploaded_file(content: bytes, filename: str) -> tuple[list, list]:
    """Parse CSV or XLSX file and return headers and rows"""
    if filename.endswith('.csv'):
        # Parse CSV
        text = content.decode('utf-8-sig')  # Handle BOM
        reader = csv.reader(StringIO(text))
        rows = list(reader)
        if not rows:
            return [], []
        headers = [h.lower().strip() for h in rows[0]]
        data_rows = rows[1:]
        return headers, data_rows
    elif filename.endswith(('.xlsx', '.xls')):
        # Parse Excel
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        headers = [cell.value.lower().strip() if cell.value else "" for cell in ws[1]]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        return headers, data_rows
    else:
        raise HTTPException(status_code=400, detail="Please upload a CSV or Excel file (.csv, .xlsx)")

# ==================== MODELS ====================

class IngredientBase(BaseModel):
    name: str
    default_unit: str = "g"
    notes: Optional[str] = None

class Ingredient(IngredientBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class IngredientPrice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ingredient_id: str
    ingredient_name: str
    store_vendor: str
    purchase_price: float
    package_size: float
    unit: str
    unit_cost: float = 0  # Calculated: purchase_price / package_size
    purchase_date: str
    is_latest: bool = True
    notes: Optional[str] = None

class PackagingBase(BaseModel):
    name: str
    store_vendor: str = ""
    purchase_price: float = 0
    package_size: float = 1
    unit_cost: float = 0
    unit: str = "piece"
    purchase_date: str = ""
    notes: Optional[str] = None

class Packaging(PackagingBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class RecipeLineItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ingredient_id: str
    ingredient_name: str
    quantity: float
    unit: str
    price_id_override: Optional[str] = None  # Override to specific price record

class PackagingLineItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    packaging_id: str
    packaging_name: str
    quantity: float

class ComponentRecipeRef(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    component_recipe_id: str
    component_name: str
    quantity: float  # Amount in grams or full batch
    use_gram_costing: bool = False

class RecipeVariant(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # e.g., "6-inch", "8-inch"
    ingredients: List[RecipeLineItem] = []
    packaging: List[PackagingLineItem] = []
    components: List[ComponentRecipeRef] = []
    prep_time_minutes: float = 0
    utility_time_minutes: float = 0
    notes: Optional[str] = None

class RecipeBase(BaseModel):
    name: str
    category: Optional[str] = None
    notes: Optional[str] = None

class Recipe(RecipeBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    variants: List[RecipeVariant] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ComponentRecipe(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: Optional[str] = None
    batch_yield_grams: float = 0  # Required for gram costing
    variants: List[RecipeVariant] = []
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "settings"
    labour_rate_per_hour: float = 10.0
    utility_rate_per_hour: float = 1.0
    currency: str = "CAD"
    currency_symbol: str = "$"

class CostBreakdown(BaseModel):
    ingredient_costs: List[Dict[str, Any]] = []
    packaging_costs: List[Dict[str, Any]] = []
    component_costs: List[Dict[str, Any]] = []
    labour_cost: float = 0
    utility_cost: float = 0
    total_ingredient_cost: float = 0
    total_packaging_cost: float = 0
    total_component_cost: float = 0
    total_cost: float = 0
    selling_price: Optional[float] = None
    profit_margin: Optional[float] = None

class Sale(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    recipe_id: str
    recipe_name: str
    variant_name: str
    sale_date: str
    customer_name: str = ""
    notes: str = ""
    selling_price: float
    total_cost: float
    labour_cost: float = 0.0
    profit: float
    profit_margin: float
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== HELPER FUNCTIONS ====================

async def get_settings() -> Settings:
    """Get or create settings document"""
    doc = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if doc:
        return Settings(**doc)
    settings = Settings()
    await db.settings.insert_one(settings.model_dump())
    return settings

async def get_ingredient_price(ingredient_id: str, price_id_override: Optional[str] = None) -> Optional[IngredientPrice]:
    """Get ingredient price - use override price ID or latest price"""
    if price_id_override:
        doc = await db.ingredient_prices.find_one(
            {"id": price_id_override},
            {"_id": 0}
        )
        if doc:
            return IngredientPrice(**doc)
    
    # Get latest price for ingredient
    doc = await db.ingredient_prices.find_one(
        {"ingredient_id": ingredient_id, "is_latest": True},
        {"_id": 0}
    )
    if doc:
        return IngredientPrice(**doc)
    
    # Fallback to any price
    doc = await db.ingredient_prices.find_one(
        {"ingredient_id": ingredient_id},
        {"_id": 0},
        sort=[("purchase_date", -1)]
    )
    if doc:
        return IngredientPrice(**doc)
    return None

async def calculate_component_cost(component_id: str, quantity: float, use_gram_costing: bool, variant_id: Optional[str] = None) -> Dict[str, Any]:
    """Calculate cost for a component recipe (uses first variant if variant_id not specified)"""
    doc = await db.component_recipes.find_one({"id": component_id}, {"_id": 0})
    if not doc:
        return {"error": f"Component {component_id} not found", "cost": 0}
    
    component = ComponentRecipe(**doc)
    settings = await get_settings()
    
    # Pick variant - use specified, or first variant, or legacy flat ingredients
    variant = None
    if component.variants:
        if variant_id:
            variant = next((v for v in component.variants if v.id == variant_id), None)
        if not variant:
            variant = component.variants[0]
    
    if not variant:
        # No variants — empty component
        return {"component_name": component.name, "quantity": quantity, "use_gram_costing": use_gram_costing, "batch_cost": 0, "cost": 0}
    
    # Calculate ingredient costs
    total_ingredient_cost = 0
    for item in variant.ingredients:
        price = await get_ingredient_price(item.ingredient_id)
        if price:
            cost = item.quantity * price.unit_cost
            total_ingredient_cost += cost
    
    # Calculate packaging costs
    total_packaging_cost = 0
    for item in variant.packaging:
        pkg = await db.packaging.find_one({"id": item.packaging_id}, {"_id": 0})
        if pkg:
            cost = item.quantity * pkg["unit_cost"]
            total_packaging_cost += cost
    
    # Labour and utilities - separate times
    prep_hours = variant.prep_time_minutes / 60
    utility_hours = variant.utility_time_minutes / 60
    labour_cost = prep_hours * settings.labour_rate_per_hour
    utility_cost = utility_hours * settings.utility_rate_per_hour
    
    batch_cost = total_ingredient_cost + total_packaging_cost + labour_cost + utility_cost
    
    if use_gram_costing and component.batch_yield_grams > 0:
        cost_per_gram = batch_cost / component.batch_yield_grams
        final_cost = cost_per_gram * quantity
    else:
        final_cost = batch_cost * quantity  # quantity as multiplier of full batch
    
    return {
        "component_name": component.name,
        "quantity": quantity,
        "use_gram_costing": use_gram_costing,
        "batch_cost": round(batch_cost, 4),
        "cost": round(final_cost, 4)
    }

# ==================== API ENDPOINTS ====================

@api_router.get("/")
async def root():
    return {"message": "Shadow Cakes Pricing Tool API"}

# ---------- SETTINGS ----------
@api_router.get("/settings", response_model=Settings)
async def get_app_settings():
    return await get_settings()

@api_router.put("/settings", response_model=Settings)
async def update_settings(settings: Settings):
    settings.id = "settings"
    await db.settings.replace_one({"id": "settings"}, settings.model_dump(), upsert=True)
    return settings

# ---------- INGREDIENTS ----------
@api_router.get("/ingredients", response_model=List[Ingredient])
async def get_ingredients():
    docs = await db.ingredients.find({}, {"_id": 0}).to_list(1000)
    return docs

@api_router.post("/ingredients", response_model=Ingredient)
async def create_ingredient(ingredient: IngredientBase):
    new_ingredient = Ingredient(**ingredient.model_dump())
    await db.ingredients.insert_one(new_ingredient.model_dump())
    return new_ingredient

@api_router.put("/ingredients/{ingredient_id}", response_model=Ingredient)
async def update_ingredient(ingredient_id: str, ingredient: IngredientBase):
    existing = await db.ingredients.find_one({"id": ingredient_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Ingredient not found")
    updated = {**existing, **ingredient.model_dump()}
    await db.ingredients.replace_one({"id": ingredient_id}, updated)
    return Ingredient(**updated)

@api_router.delete("/ingredients/{ingredient_id}")
async def delete_ingredient(ingredient_id: str):
    result = await db.ingredients.delete_one({"id": ingredient_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ingredient not found")
    # Also delete associated prices
    await db.ingredient_prices.delete_many({"ingredient_id": ingredient_id})
    return {"status": "deleted"}

# ---------- INGREDIENT PRICES ----------
@api_router.get("/ingredient-prices", response_model=List[IngredientPrice])
async def get_ingredient_prices():
    docs = await db.ingredient_prices.find({}, {"_id": 0}).to_list(5000)
    return docs

@api_router.get("/ingredient-prices/{ingredient_id}", response_model=List[IngredientPrice])
async def get_prices_for_ingredient(ingredient_id: str):
    docs = await db.ingredient_prices.find({"ingredient_id": ingredient_id}, {"_id": 0}).to_list(100)
    return docs

@api_router.post("/ingredient-prices", response_model=IngredientPrice)
async def create_ingredient_price(price: IngredientPrice):
    # Calculate unit cost
    price.unit_cost = price.purchase_price / price.package_size if price.package_size > 0 else 0
    # Set other prices for this ingredient as not latest
    if price.is_latest:
        await db.ingredient_prices.update_many(
            {"ingredient_id": price.ingredient_id},
            {"$set": {"is_latest": False}}
        )
    await db.ingredient_prices.insert_one(price.model_dump())
    return price

@api_router.delete("/ingredient-prices/{price_id}")
async def delete_ingredient_price(price_id: str):
    result = await db.ingredient_prices.delete_one({"id": price_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Price not found")
    return {"status": "deleted"}

@api_router.put("/ingredient-prices/{price_id}")
async def update_ingredient_price(price_id: str, price: IngredientPrice):
    existing = await db.ingredient_prices.find_one({"id": price_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Price not found")
    update_data = price.model_dump()
    update_data["id"] = price_id
    if update_data["package_size"] > 0:
        update_data["unit_cost"] = update_data["purchase_price"] / update_data["package_size"]
    await db.ingredient_prices.replace_one({"id": price_id}, update_data)
    return {**update_data}

# ---------- PACKAGING ----------
@api_router.get("/packaging", response_model=List[Packaging])
async def get_packaging():
    docs = await db.packaging.find({}, {"_id": 0}).to_list(1000)
    return docs

@api_router.post("/packaging", response_model=Packaging)
async def create_packaging(packaging: PackagingBase):
    data = packaging.model_dump()
    if data["package_size"] > 0 and data["purchase_price"] > 0 and data["unit_cost"] == 0:
        data["unit_cost"] = data["purchase_price"] / data["package_size"]
    new_packaging = Packaging(**data)
    await db.packaging.insert_one(new_packaging.model_dump())
    return new_packaging

@api_router.put("/packaging/{packaging_id}", response_model=Packaging)
async def update_packaging(packaging_id: str, packaging: PackagingBase):
    existing = await db.packaging.find_one({"id": packaging_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Packaging not found")
    data = packaging.model_dump()
    if data["package_size"] > 0 and data["purchase_price"] > 0:
        data["unit_cost"] = data["purchase_price"] / data["package_size"]
    updated = {**existing, **data}
    await db.packaging.replace_one({"id": packaging_id}, updated)
    return Packaging(**updated)

@api_router.delete("/packaging/{packaging_id}")
async def delete_packaging(packaging_id: str):
    result = await db.packaging.delete_one({"id": packaging_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Packaging not found")
    return {"status": "deleted"}

# ---------- COMPONENT RECIPES ----------
@api_router.get("/component-recipes", response_model=List[ComponentRecipe])
async def get_component_recipes():
    docs = await db.component_recipes.find({}, {"_id": 0}).to_list(1000)
    return docs

@api_router.get("/component-recipes/{component_id}", response_model=ComponentRecipe)
async def get_component_recipe(component_id: str):
    doc = await db.component_recipes.find_one({"id": component_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Component recipe not found")
    return doc

@api_router.post("/component-recipes", response_model=ComponentRecipe)
async def create_component_recipe(component: ComponentRecipe):
    await db.component_recipes.insert_one(component.model_dump())
    return component

@api_router.put("/component-recipes/{component_id}", response_model=ComponentRecipe)
async def update_component_recipe(component_id: str, component: ComponentRecipe):
    component.id = component_id
    result = await db.component_recipes.replace_one({"id": component_id}, component.model_dump())
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Component recipe not found")
    return component

@api_router.delete("/component-recipes/{component_id}")
async def delete_component_recipe(component_id: str):
    result = await db.component_recipes.delete_one({"id": component_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Component recipe not found")
    return {"status": "deleted"}

# ---------- RECIPES ----------
@api_router.get("/recipes", response_model=List[Recipe])
async def get_recipes():
    docs = await db.recipes.find({}, {"_id": 0}).to_list(1000)
    return docs

@api_router.get("/recipes/{recipe_id}", response_model=Recipe)
async def get_recipe(recipe_id: str):
    doc = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return doc

@api_router.post("/recipes", response_model=Recipe)
async def create_recipe(recipe: Recipe):
    await db.recipes.insert_one(recipe.model_dump())
    return recipe

@api_router.put("/recipes/{recipe_id}", response_model=Recipe)
async def update_recipe(recipe_id: str, recipe: Recipe):
    recipe.id = recipe_id
    result = await db.recipes.replace_one({"id": recipe_id}, recipe.model_dump())
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return recipe

@api_router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str):
    result = await db.recipes.delete_one({"id": recipe_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return {"status": "deleted"}

# ---------- COST CALCULATION ----------
@api_router.get("/recipes/{recipe_id}/variants/{variant_id}/cost")
async def calculate_variant_cost(recipe_id: str, variant_id: str, selling_price: Optional[float] = None):
    """Calculate full cost breakdown for a recipe variant"""
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    
    variant = None
    for v in recipe.get("variants", []):
        if v["id"] == variant_id:
            variant = RecipeVariant(**v)
            break
    
    if not variant:
        raise HTTPException(status_code=404, detail="Variant not found")
    
    settings = await get_settings()
    breakdown = CostBreakdown()
    
    # Calculate ingredient costs
    for item in variant.ingredients:
        price = await get_ingredient_price(item.ingredient_id, item.price_id_override)
        if price:
            cost = item.quantity * price.unit_cost
            breakdown.ingredient_costs.append({
                "ingredient_name": item.ingredient_name,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_cost": round(price.unit_cost, 4),
                "store_vendor": price.store_vendor,
                "brand": price.notes,  # notes field stores brand
                "is_override": item.price_id_override is not None,
                "cost": round(cost, 4)
            })
            breakdown.total_ingredient_cost += cost
    
    # Calculate packaging costs
    for item in variant.packaging:
        pkg = await db.packaging.find_one({"id": item.packaging_id}, {"_id": 0})
        if pkg:
            cost = item.quantity * pkg["unit_cost"]
            breakdown.packaging_costs.append({
                "packaging_name": item.packaging_name,
                "quantity": item.quantity,
                "unit_cost": pkg["unit_cost"],
                "cost": round(cost, 4)
            })
            breakdown.total_packaging_cost += cost
    
    # Calculate component costs
    for comp_ref in variant.components:
        comp_cost = await calculate_component_cost(
            comp_ref.component_recipe_id,
            comp_ref.quantity,
            comp_ref.use_gram_costing
        )
        breakdown.component_costs.append(comp_cost)
        breakdown.total_component_cost += comp_cost.get("cost", 0)
    
    # Calculate labour and utility costs (separate times)
    prep_hours = variant.prep_time_minutes / 60
    utility_hours = variant.utility_time_minutes / 60
    breakdown.labour_cost = round(prep_hours * settings.labour_rate_per_hour, 4)
    breakdown.utility_cost = round(utility_hours * settings.utility_rate_per_hour, 4)
    
    # Total cost
    breakdown.total_ingredient_cost = round(breakdown.total_ingredient_cost, 2)
    breakdown.total_packaging_cost = round(breakdown.total_packaging_cost, 2)
    breakdown.total_component_cost = round(breakdown.total_component_cost, 2)
    breakdown.total_cost = round(
        breakdown.total_ingredient_cost +
        breakdown.total_packaging_cost +
        breakdown.total_component_cost +
        breakdown.labour_cost +
        breakdown.utility_cost,
        2
    )
    
    # Calculate margin if selling price provided
    if selling_price is not None and selling_price > 0:
        breakdown.selling_price = selling_price
        profit = selling_price - breakdown.total_cost
        breakdown.profit_margin = round((profit / selling_price) * 100, 2)
    
    return {
        "recipe_name": recipe["name"],
        "variant_name": variant.name,
        "prep_time_minutes": variant.prep_time_minutes,
        "utility_time_minutes": variant.utility_time_minutes,
        "settings": settings.model_dump(),
        "breakdown": breakdown.model_dump()
    }

# ---------- EXCEL IMPORT ----------
@api_router.post("/import/ingredients")
async def import_ingredient_pricing(file: UploadFile = File(...)):
    """
    Import ingredient pricing from CSV or Excel.
    Expected columns: ingredient_name, store_vendor, purchase_price, package_size, unit, purchase_date, brand
    """
    content = await file.read()
    headers, data_rows = parse_uploaded_file(content, file.filename)
    
    required = ["ingredient_name", "store_vendor", "purchase_price", "package_size", "unit"]
    missing = [r for r in required if r not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")
    
    # Clear existing data (replace behavior)
    await db.ingredient_prices.delete_many({})
    await db.ingredients.delete_many({})
    
    ingredients_map = {}  # name -> id
    prices_added = 0
    errors = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        try:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            
            ingredient_name = str(row_dict.get("ingredient_name", "") or "").strip()
            if not ingredient_name or ingredient_name.lower() == "none":
                continue
            
            # Create or get ingredient
            if ingredient_name not in ingredients_map:
                ingredient = Ingredient(name=ingredient_name, default_unit=str(row_dict.get("unit", "g") or "g"))
                await db.ingredients.insert_one(ingredient.model_dump())
                ingredients_map[ingredient_name] = ingredient.id
            
            ingredient_id = ingredients_map[ingredient_name]
            
            # Create price record
            purchase_price = float(row_dict.get("purchase_price", 0) or 0)
            package_size = float(row_dict.get("package_size", 1) or 1)
            unit_cost = purchase_price / package_size if package_size > 0 else 0
            
            purchase_date = row_dict.get("purchase_date", "")
            if isinstance(purchase_date, datetime):
                purchase_date = purchase_date.isoformat()
            else:
                purchase_date = str(purchase_date) if purchase_date else datetime.now(timezone.utc).isoformat()
            
            price = IngredientPrice(
                ingredient_id=ingredient_id,
                ingredient_name=ingredient_name,
                store_vendor=str(row_dict.get("store_vendor", "Default") or "Default"),
                purchase_price=purchase_price,
                package_size=package_size,
                unit=str(row_dict.get("unit", "g") or "g"),
                unit_cost=unit_cost,
                purchase_date=purchase_date,
                is_latest=True,
                notes=str(row_dict.get("brand", "") or "")
            )
            await db.ingredient_prices.insert_one(price.model_dump())
            prices_added += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {
        "status": "success",
        "ingredients_created": len(ingredients_map),
        "prices_added": prices_added,
        "errors": errors[:10] if errors else []
    }

@api_router.post("/import/recipes")
async def import_recipes(file: UploadFile = File(...)):
    """
    Import recipes from CSV or Excel.
    New format: recipe_name, variant_name, line_type, item_name, quantity, unit, prep_time_minutes, utility_time_minutes, category, notes
    Also supports legacy format: recipe_name, variant_name, ingredient_name, quantity, unit, ...
    line_type = ingredient | packaging | component
    """
    content = await file.read()
    headers, data_rows = parse_uploaded_file(content, file.filename)
    
    # Detect format: new (has line_type + item_name) or legacy (has ingredient_name)
    has_line_type = "line_type" in headers and "item_name" in headers
    has_legacy = "ingredient_name" in headers
    
    if not has_line_type and not has_legacy:
        raise HTTPException(status_code=400, detail="Missing required columns. Expected either (line_type, item_name) or (ingredient_name)")
    
    if "recipe_name" not in headers or "variant_name" not in headers:
        raise HTTPException(status_code=400, detail="Missing required columns: recipe_name, variant_name")
    
    # Clear existing recipes (replace behavior)
    await db.recipes.delete_many({})
    
    recipes_map = {}
    ingredients_list = await db.ingredients.find({}, {"_id": 0}).to_list(5000)
    ingredients_lookup = {i["name"].lower(): i for i in ingredients_list}
    packaging_list = await db.packaging.find({}, {"_id": 0}).to_list(5000)
    packaging_lookup = {p["name"].lower(): p for p in packaging_list}
    components_list = await db.component_recipes.find({}, {"_id": 0}).to_list(5000)
    # Build component lookup: "name — variant" -> (comp_id, variant_id)
    component_lookup = {}
    for comp in components_list:
        for var in comp.get("variants", []):
            key = f"{comp['name']} — {var['name']}".lower()
            component_lookup[key] = {"comp_id": comp["id"], "comp_name": f"{comp['name']} — {var['name']}"}
        # Also just by name (picks first variant)
        if comp.get("variants"):
            component_lookup[comp["name"].lower()] = {"comp_id": comp["id"], "comp_name": comp["name"]}
    
    errors = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        try:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            
            recipe_name = str(row_dict.get("recipe_name", "") or "").strip()
            variant_name = str(row_dict.get("variant_name", "") or "").strip()
            
            if not recipe_name or not variant_name:
                continue
            
            # Determine line_type and item_name
            if has_line_type:
                line_type = str(row_dict.get("line_type", "ingredient") or "ingredient").strip().lower()
                item_name = str(row_dict.get("item_name", "") or "").strip()
            else:
                line_type = "ingredient"
                item_name = str(row_dict.get("ingredient_name", "") or "").strip()
            
            # Get or create recipe
            if recipe_name not in recipes_map:
                recipes_map[recipe_name] = Recipe(
                    name=recipe_name,
                    category=str(row_dict.get("category", "") or ""),
                    notes=str(row_dict.get("notes", "") or ""),
                    variants=[]
                )
            
            recipe = recipes_map[recipe_name]
            
            # Get or create variant
            variant = next((v for v in recipe.variants if v.name == variant_name), None)
            if not variant:
                prep_time = float(row_dict.get("prep_time_minutes", 0) or 0)
                utility_time = float(row_dict.get("utility_time_minutes", 0) or 0)
                variant = RecipeVariant(
                    name=variant_name,
                    prep_time_minutes=prep_time,
                    utility_time_minutes=utility_time,
                    notes=str(row_dict.get("notes", "") or "")
                )
                recipe.variants.append(variant)
            
            if not item_name:
                continue
            
            if line_type == "ingredient":
                ingredient_key = item_name.lower()
                ingredient_id = ingredients_lookup.get(ingredient_key, {}).get("id", "")
                if not ingredient_id:
                    new_ing = Ingredient(name=item_name, default_unit=str(row_dict.get("unit", "g") or "g"))
                    await db.ingredients.insert_one(new_ing.model_dump())
                    ingredients_lookup[ingredient_key] = new_ing.model_dump()
                    ingredient_id = new_ing.id
                
                variant.ingredients.append(RecipeLineItem(
                    ingredient_id=ingredient_id,
                    ingredient_name=item_name,
                    quantity=float(row_dict.get("quantity", 0) or 0),
                    unit=str(row_dict.get("unit", "g") or "g")
                ))
            
            elif line_type == "packaging":
                pkg_key = item_name.lower()
                pkg = packaging_lookup.get(pkg_key)
                if pkg:
                    variant.packaging.append(PackagingLineItem(
                        packaging_id=pkg["id"],
                        packaging_name=item_name,
                        quantity=float(row_dict.get("quantity", 1) or 1)
                    ))
                else:
                    errors.append(f"Row {row_idx}: Packaging '{item_name}' not found")
            
            elif line_type == "component":
                comp_key = item_name.lower()
                comp_info = component_lookup.get(comp_key)
                if comp_info:
                    variant.components.append(ComponentRecipeRef(
                        component_recipe_id=comp_info["comp_id"],
                        component_name=comp_info["comp_name"],
                        quantity=1,
                        use_gram_costing=False
                    ))
                else:
                    errors.append(f"Row {row_idx}: Component '{item_name}' not found")
                
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    for recipe in recipes_map.values():
        await db.recipes.insert_one(recipe.model_dump())
    
    total_variants = sum(len(r.variants) for r in recipes_map.values())
    
    return {
        "status": "success",
        "recipes_created": len(recipes_map),
        "variants_created": total_variants,
        "errors": errors[:10] if errors else []
    }

@api_router.post("/import/packaging")
async def import_packaging(file: UploadFile = File(...)):
    """
    Import packaging from CSV or Excel (same format as ingredients).
    Expected columns: packaging_name, store_vendor, purchase_price, package_size, unit, purchase_date, notes
    """
    content = await file.read()
    headers, data_rows = parse_uploaded_file(content, file.filename)
    
    required = ["packaging_name", "store_vendor", "purchase_price", "package_size", "unit"]
    missing = [r for r in required if r not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")
    
    # Clear existing packaging (replace behavior)
    await db.packaging.delete_many({})
    
    items_added = 0
    errors = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        try:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            
            name = str(row_dict.get("packaging_name", "") or "").strip()
            if not name or name.lower() == "none":
                continue
            
            # Calculate unit cost from purchase_price / package_size
            purchase_price = float(row_dict.get("purchase_price", 0) or 0)
            package_size = float(row_dict.get("package_size", 1) or 1)
            unit_cost = purchase_price / package_size if package_size > 0 else purchase_price
            
            purchase_date = row_dict.get("purchase_date", "")
            if isinstance(purchase_date, datetime):
                purchase_date = purchase_date.isoformat()
            else:
                purchase_date = str(purchase_date) if purchase_date else ""
            
            packaging = Packaging(
                name=name,
                store_vendor=str(row_dict.get("store_vendor", "") or ""),
                purchase_price=purchase_price,
                package_size=package_size,
                unit_cost=unit_cost,
                unit=str(row_dict.get("unit", "piece") or "piece"),
                purchase_date=purchase_date,
                notes=str(row_dict.get("notes", "") or "")
            )
            await db.packaging.insert_one(packaging.model_dump())
            items_added += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {
        "status": "success",
        "packaging_added": items_added,
        "errors": errors[:10] if errors else []
    }

# ---------- BULK DELETE ENDPOINTS ----------
class BulkDeleteRequest(BaseModel):
    ids: List[str]

@api_router.post("/ingredients/bulk-delete")
async def bulk_delete_ingredients(request: BulkDeleteRequest):
    """Delete multiple ingredients and their prices"""
    deleted_count = 0
    for ing_id in request.ids:
        result = await db.ingredients.delete_one({"id": ing_id})
        if result.deleted_count > 0:
            deleted_count += 1
            await db.ingredient_prices.delete_many({"ingredient_id": ing_id})
    return {"status": "success", "deleted_count": deleted_count}

@api_router.post("/packaging/bulk-delete")
async def bulk_delete_packaging(request: BulkDeleteRequest):
    """Delete multiple packaging items"""
    deleted_count = 0
    for pkg_id in request.ids:
        result = await db.packaging.delete_one({"id": pkg_id})
        if result.deleted_count > 0:
            deleted_count += 1
    return {"status": "success", "deleted_count": deleted_count}

@api_router.post("/component-recipes/bulk-delete")
async def bulk_delete_components(request: BulkDeleteRequest):
    """Delete multiple component recipes"""
    deleted_count = 0
    for comp_id in request.ids:
        result = await db.component_recipes.delete_one({"id": comp_id})
        if result.deleted_count > 0:
            deleted_count += 1
    return {"status": "success", "deleted_count": deleted_count}

# ---------- SALES ----------
@api_router.get("/sales/summary")
async def get_sales_summary():
    """Get aggregated sales summary for dashboard"""
    docs = await db.sales.find({}, {"_id": 0}).to_list(5000)
    total_sales = len(docs)
    total_revenue = sum(d.get("selling_price", 0) for d in docs)
    total_cost = sum(d.get("total_cost", 0) for d in docs)
    total_labour = sum(d.get("labour_cost", 0) for d in docs)
    total_profit = sum(d.get("profit", 0) for d in docs)
    return {
        "total_sales": total_sales,
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "total_hourly_pay": total_labour,
        "total_profit": total_profit
    }

@api_router.get("/sales", response_model=List[Sale])
async def get_sales():
    docs = await db.sales.find({}, {"_id": 0}).sort("sale_date", -1).to_list(5000)
    return docs

@api_router.post("/sales", response_model=Sale)
async def create_sale(sale: Sale):
    await db.sales.insert_one(sale.model_dump())
    return sale

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str):
    result = await db.sales.delete_one({"id": sale_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sale not found")
    return {"status": "deleted"}

@api_router.put("/sales/{sale_id}")
async def update_sale(sale_id: str, sale: Sale):
    update_data = sale.model_dump()
    update_data.pop("id", None)
    result = await db.sales.update_one({"id": sale_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sale not found")
    return {"status": "updated"}

@api_router.post("/sales/bulk-delete")
async def bulk_delete_sales(request: BulkDeleteRequest):
    """Delete multiple sales"""
    deleted_count = 0
    for sale_id in request.ids:
        result = await db.sales.delete_one({"id": sale_id})
        if result.deleted_count > 0:
            deleted_count += 1
    return {"status": "success", "deleted_count": deleted_count}

@api_router.post("/import/sales")
async def import_sales(file: UploadFile = File(...)):
    """
    Import sales from CSV or Excel.
    Uses the exact same format as the export:
    Sale Date, Recipe, Variant, Customer, Selling Price, Total Cost, Labour Cost, Profit, Profit Margin, Notes
    """
    content = await file.read()
    headers, data_rows = parse_uploaded_file(content, file.filename)
    
    # Map exported header names to internal field names
    header_map = {
        "sale date": "sale_date",
        "recipe": "recipe_name",
        "variant": "variant_name",
        "customer": "customer_name",
        "selling price": "selling_price",
        "total cost": "total_cost",
        "labour cost": "labour_cost",
        "profit": "profit",
        "profit margin": "profit_margin",
        "notes": "notes",
    }
    
    # Normalize headers (lowercase for matching)
    normalized = [header_map.get(h.strip().lower(), h.strip().lower()) for h in headers]
    
    required = ["sale_date", "recipe_name", "selling_price", "total_cost"]
    missing = [r for r in required if r not in normalized]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}. Expected headers: Sale Date, Recipe, Variant, Customer, Selling Price, Total Cost, Labour Cost, Profit, Profit Margin, Notes")
    
    sales_created = 0
    errors = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        try:
            row_dict = {normalized[i]: row[i] if i < len(row) else None for i in range(len(normalized))}
            
            sale_date = row_dict.get("sale_date", "")
            if isinstance(sale_date, datetime):
                sale_date = sale_date.strftime("%Y-%m-%d")
            else:
                sale_date = str(sale_date or "").strip()
            
            if not sale_date:
                continue
            
            selling_price = float(str(row_dict.get("selling_price", 0) or 0).replace("$", "").replace(",", ""))
            total_cost = float(str(row_dict.get("total_cost", 0) or 0).replace("$", "").replace(",", ""))
            labour_cost = float(str(row_dict.get("labour_cost", 0) or 0).replace("$", "").replace(",", ""))
            
            # Parse profit — calculate if not provided
            profit_raw = str(row_dict.get("profit", "") or "").replace("$", "").replace(",", "")
            profit = float(profit_raw) if profit_raw else selling_price - total_cost
            
            # Parse margin — calculate if not provided
            margin_raw = str(row_dict.get("profit_margin", "") or "").replace("%", "").strip()
            if margin_raw:
                profit_margin = float(margin_raw)
            else:
                profit_margin = round((profit / selling_price) * 100, 2) if selling_price > 0 else 0
            
            sale = Sale(
                recipe_id="",
                recipe_name=str(row_dict.get("recipe_name", "") or "").strip(),
                variant_name=str(row_dict.get("variant_name", "") or "").strip(),
                sale_date=sale_date,
                customer_name=str(row_dict.get("customer_name", "") or "").strip(),
                notes=str(row_dict.get("notes", "") or "").strip(),
                selling_price=selling_price,
                total_cost=total_cost,
                labour_cost=labour_cost,
                profit=profit,
                profit_margin=profit_margin,
            )
            await db.sales.insert_one(sale.model_dump())
            sales_created += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {
        "status": "success",
        "sales_imported": sales_created,
        "errors": errors[:10] if errors else []
    }

@api_router.post("/import/components")
async def import_component_recipes(file: UploadFile = File(...)):
    """
    Import component recipes from CSV or Excel.
    New format: recipe_name, variant_name, line_type, item_name, quantity, unit, prep_time_minutes, utility_time_minutes, category, notes
    Also supports legacy format with ingredient_name column.
    """
    content = await file.read()
    headers, data_rows = parse_uploaded_file(content, file.filename)
    
    # Detect format
    has_line_type = "line_type" in headers and "item_name" in headers
    has_legacy_ingredient = "ingredient_name" in headers
    
    # Support component_name or recipe_name
    has_component_name = "component_name" in headers
    has_recipe_name = "recipe_name" in headers
    if not has_component_name and not has_recipe_name:
        raise HTTPException(status_code=400, detail="Missing required column: component_name or recipe_name")
    
    name_col = "component_name" if has_component_name else "recipe_name"
    has_variant = "variant_name" in headers
    
    if not has_line_type and not has_legacy_ingredient:
        raise HTTPException(status_code=400, detail="Missing required columns. Expected either (line_type, item_name) or (ingredient_name)")
    
    await db.component_recipes.delete_many({})
    
    components_map = {}
    variant_map = {}
    ingredients_list = await db.ingredients.find({}, {"_id": 0}).to_list(5000)
    ingredients_lookup = {i["name"].lower(): i for i in ingredients_list}
    packaging_list = await db.packaging.find({}, {"_id": 0}).to_list(5000)
    packaging_lookup = {p["name"].lower(): p for p in packaging_list}
    
    errors = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        try:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            
            component_name = str(row_dict.get(name_col, "") or "").strip()
            variant_name = str(row_dict.get("variant_name", "Default") or "Default").strip() if has_variant else "Default"
            
            if not component_name or component_name.lower() == "none":
                continue
            
            # Determine line_type and item_name
            if has_line_type:
                line_type = str(row_dict.get("line_type", "ingredient") or "ingredient").strip().lower()
                item_name = str(row_dict.get("item_name", "") or "").strip()
            else:
                line_type = "ingredient"
                item_name = str(row_dict.get("ingredient_name", "") or "").strip()
            
            if component_name not in components_map:
                components_map[component_name] = ComponentRecipe(
                    name=component_name,
                    category=str(row_dict.get("category", "") or ""),
                    batch_yield_grams=float(row_dict.get("batch_yield_grams", 0) or 0),
                    notes=str(row_dict.get("notes", "") or ""),
                    variants=[]
                )
            
            component = components_map[component_name]
            key = (component_name, variant_name)
            
            if key not in variant_map:
                prep_time = float(row_dict.get("prep_time_minutes", 0) or 0)
                utility_time = float(row_dict.get("utility_time_minutes", 0) or 0)
                variant = RecipeVariant(
                    name=variant_name,
                    prep_time_minutes=prep_time,
                    utility_time_minutes=utility_time,
                    notes=str(row_dict.get("notes", "") or "")
                )
                component.variants.append(variant)
                variant_map[key] = variant
            
            variant = variant_map[key]
            
            if not item_name:
                continue
            
            if line_type == "ingredient":
                ingredient_key = item_name.lower()
                ingredient_id = ingredients_lookup.get(ingredient_key, {}).get("id", "")
                if not ingredient_id:
                    new_ing = Ingredient(name=item_name, default_unit=str(row_dict.get("unit", "g") or "g"))
                    await db.ingredients.insert_one(new_ing.model_dump())
                    ingredients_lookup[ingredient_key] = new_ing.model_dump()
                    ingredient_id = new_ing.id
                
                variant.ingredients.append(RecipeLineItem(
                    ingredient_id=ingredient_id,
                    ingredient_name=item_name,
                    quantity=float(row_dict.get("quantity", 0) or 0),
                    unit=str(row_dict.get("unit", "g") or "g")
                ))
            
            elif line_type == "packaging":
                pkg = packaging_lookup.get(item_name.lower())
                if pkg:
                    variant.packaging.append(PackagingLineItem(
                        packaging_id=pkg["id"],
                        packaging_name=item_name,
                        quantity=float(row_dict.get("quantity", 1) or 1)
                    ))
                else:
                    errors.append(f"Row {row_idx}: Packaging '{item_name}' not found")
                
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    for component in components_map.values():
        await db.component_recipes.insert_one(component.model_dump())
    
    total_variants = sum(len(c.variants) for c in components_map.values())
    
    return {
        "status": "success",
        "components_created": len(components_map),
        "variants_created": total_variants,
        "errors": errors[:10] if errors else []
    }

# ---------- COMPONENT COST CALCULATION ----------
@api_router.get("/component-recipes/{component_id}/variants/{variant_id}/cost")
async def calculate_component_variant_cost(component_id: str, variant_id: str, selling_price: Optional[float] = None):
    """Calculate full cost breakdown for a component variant (same as recipe cost)"""
    doc = await db.component_recipes.find_one({"id": component_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Component not found")
    
    component = ComponentRecipe(**doc)
    variant = next((v for v in component.variants if v.id == variant_id), None)
    if not variant:
        raise HTTPException(status_code=404, detail="Variant not found")
    
    settings = await get_settings()
    breakdown = CostBreakdown()
    
    for item in variant.ingredients:
        price = await get_ingredient_price(item.ingredient_id, item.price_id_override)
        if price:
            cost = item.quantity * price.unit_cost
            breakdown.ingredient_costs.append({
                "ingredient_name": item.ingredient_name,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_cost": round(price.unit_cost, 4),
                "store_vendor": price.store_vendor,
                "brand": price.notes,
                "is_override": item.price_id_override is not None,
                "cost": round(cost, 4)
            })
            breakdown.total_ingredient_cost += cost
    
    for item in variant.packaging:
        pkg = await db.packaging.find_one({"id": item.packaging_id}, {"_id": 0})
        if pkg:
            cost = item.quantity * pkg["unit_cost"]
            breakdown.packaging_costs.append({
                "packaging_name": item.packaging_name,
                "quantity": item.quantity,
                "unit_cost": pkg["unit_cost"],
                "cost": round(cost, 4)
            })
            breakdown.total_packaging_cost += cost
    
    for comp_ref in variant.components:
        comp_cost = await calculate_component_cost(
            comp_ref.component_recipe_id,
            comp_ref.quantity,
            comp_ref.use_gram_costing
        )
        breakdown.component_costs.append(comp_cost)
        breakdown.total_component_cost += comp_cost.get("cost", 0)
    
    prep_hours = variant.prep_time_minutes / 60
    utility_hours = variant.utility_time_minutes / 60
    breakdown.labour_cost = round(prep_hours * settings.labour_rate_per_hour, 4)
    breakdown.utility_cost = round(utility_hours * settings.utility_rate_per_hour, 4)
    
    breakdown.total_ingredient_cost = round(breakdown.total_ingredient_cost, 2)
    breakdown.total_packaging_cost = round(breakdown.total_packaging_cost, 2)
    breakdown.total_component_cost = round(breakdown.total_component_cost, 2)
    breakdown.total_cost = round(
        breakdown.total_ingredient_cost + breakdown.total_packaging_cost +
        breakdown.total_component_cost + breakdown.labour_cost + breakdown.utility_cost, 2
    )
    
    if selling_price is not None and selling_price > 0:
        breakdown.selling_price = selling_price
        profit = selling_price - breakdown.total_cost
        breakdown.profit_margin = round((profit / selling_price) * 100, 2)
    
    return {
        "recipe_name": component.name,
        "variant_name": variant.name,
        "prep_time_minutes": variant.prep_time_minutes,
        "utility_time_minutes": variant.utility_time_minutes,
        "batch_yield_grams": component.batch_yield_grams,
        "settings": settings.model_dump(),
        "breakdown": breakdown.model_dump()
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
